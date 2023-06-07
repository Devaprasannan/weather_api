import urllib.request as url
import json
import requests
import pandas as pd
import csv
import warnings
warnings.filterwarnings("ignore")
#step1
#Creating a class, requesting the API, and creating function for creating the dataframe,printing and writing the dataframe in CSV format.
class weather:
  def createDf(self):
      self.df = pd.DataFrame()
  def printDf(self):
      print(self.df)
  def writeDF(self):
      self.df.to_csv('weather_output.csv', mode='a', index=False, header=True)
  def pop_data(self,xx):
   base_url = "https://api.openweathermap.org/data/2.5/weather?"
   api_key = '8930feff9a4ae2eb052c7055ec2b99f1'
   complete_url = base_url + "appid=" + api_key + "&q=" +xx
   response=requests.get(complete_url)
   self.data=response.json()
   print(self.data)
   self.df=self.df.append({'lon': self.data.get('coord').get('lon'), 'lat': self.data.get('coord').get('lat'),
              'weather_id': self.data.get('weather')[0].get('id'),'description':self.data.get('weather')[0].get('description'),'icon':self.data.get('weather')[0].get('icon'),
              'base':self.data.get('base'),'temp':self.data.get('main').get('temp'),'feels_like':self.data.get('main').get('feels_like'),'temp_min':self.data.get('main').get('temp_min'),
              'temp_max': self.data.get('main').get('temp_max'), 'pressure': self.data.get('main').get('pressure'),'humidity': self.data.get('main').get('humidity'),
              'visibility':self.data.get( 'visibility'),'speed':self.data.get('wind').get('speed'),'deg':self.data.get('wind').get('deg'),'all':self.data.get('clouds').get('all'),
              'dt': self.data.get('dt'), 'type': self.data.get('sys').get('type'), 'sys_id': self.data.get('sys').get('id'),'country': self.data.get('sys').get('country'), 'sunrise': self.data.get('sys').get('sunrise'),
              'sunset':self.data.get('sys').get('sunset'),'timezone':self.data.get('timezone'),'id':self.data.get('id'),'name':self.data.get('name'),'cod':self.data.get('cod')}, ignore_index=True)
from openpyxl import Workbook, load_workbook
a1 = load_workbook(filename='10 cities.xlsx')
sh = a1['Sheet2']
row_ct = sh.max_row
col_ct = sh.max_column
y = []
obj = weather()
obj.createDf()
for i in range(2, row_ct + 1):
    z = (sh.cell(row=i, column=1).value)
    y.append(z)
for xx in range(len(y)):
     obj.pop_data(y[xx])
obj.printDf()
obj.writeDF()
#step2
#creating the class for visualisation, using the URL link of weather data stored in github we are creating the various charts
import pandas as pd
import seaborn as sns
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.pyplot import figure
import datetime
from datetime import datetime
df1 = pd.read_csv('https://raw.githubusercontent.com/Devaprasannan/weather1/main/file1.csv')
print(df1.head(5))
class visual:
    def ca(self):
        self.df1 = pd.read_csv('https://raw.githubusercontent.com/Devaprasannan/weather1/main/file1.csv')
        print(self.df1.head(5))
    def bar(self):
        import numpy as np
        import seaborn as sns
        import matplotlib.pyplot as plt
        fig = plt.figure()
        plt.figure(figsize=(15, 10))
        result = self.df1.groupby(["country"])['humidity'].aggregate(np.median).reset_index().sort_values('humidity')
        sns.barplot(x='country', y="humidity", data=self.df1, order=result['country'], palette='flare')
        plt.title('High Humidity Data')
        plt.xlabel('country', fontsize=50)
        plt.ylabel('Humidity', fontsize=40)
        plt.show()
    def bar1(self):
        fig = plt.figure()
        plt.figure(figsize=(15, 10))
        result = self.df1.groupby(["country"])['speed'].aggregate(np.median).reset_index().sort_values('speed')
        sns.barplot(x='country', y="speed", data=self.df1, order=result['country'], palette='flare')
        plt.title('Wind_speed Data')
        plt.xlabel('country', fontsize=50)
        plt.ylabel('Wind_speed', fontsize=40)
        plt.show()
    def bar2(self):

        fig = plt.figure()
        plt.figure(figsize=(15, 10))
        result = self.df1.groupby(["country"])['pressure'].aggregate(np.median).reset_index().sort_values('pressure')
        sns.barplot(x='country', y="pressure", data=self.df1, order=result['country'], palette='flare')
        plt.title('Wind Pressure Data')
        plt.xlabel('country', fontsize=50)
        plt.ylabel('Wind_pressure', fontsize=40)
        plt.show()
    def heat(self):
        sns.set_theme()
        plt.title('Countries and cities with respect to Humidity')
        df2 = self.df1.pivot("country", "name", "humidity")
        f, ax = plt.subplots(figsize=(20, 4))
        sns.heatmap(df2, annot=True, fmt=".0f", linewidths=0.1, ax=ax)
        plt.show()
    def density_map(self):
        sns.set_theme(style="whitegrid")
        sns.displot(
            data=self.df1,
            x="temp", hue="country",
            kind="kde", height=6,
            multiple="fill", clip=(0, None),
            palette="ch:rot=-.25,hue=1,light=.75")
        plt.title('Temperature density of countries')
        plt.show()
    def precipitaion(self):
        df3 = pd.read_csv('https://raw.githubusercontent.com/Devaprasannan/weather1/main/pr_timeseries_annual_cru_1901-2021_IND.csv')
        df3.head(5)
        import seaborn as sns
        sns.set_theme(style="darkgrid")
        plt.figure(figsize=(15, 10))
        sns.lineplot(x="Year", y="India", data=df3)
        plt.title('Precipitation data from 1901-2021')
        plt.show()
    def lineplot(self):
        df2 = pd.read_csv('https://raw.githubusercontent.com/Devaprasannan/weather1/main/pr_climatology_annual-seasonal_cru_1991-2020_IND.csv')
        df2.head(5)
        sns.set_theme(style="whitegrid")
        plt.figure(figsize=(15, 10))
        sns.lineplot(data=df2, palette="tab10", linewidth=2.5)
        plt.title('Annual rainfal seasonal Data')
        plt.show()
c=visual()
c.ca()
c.bar()
c.bar1()
c.bar2()
c.heat()
c.density_map()
c.precipitaion()
c.lineplot()
#step3
#using same link available in github the database is created in postgreSQL and various quries are excecuted
import psycopg2
import pandas as pd
from matplotlib.pyplot import figure
import datetime
from datetime import datetime
from sqlalchemy import create_engine
class etl:
 def get_data(self):
        self.df1=pd.read_csv('https://raw.githubusercontent.com/Devaprasannan/weather1/main/file22.csv')
        self.df1.head(5)
 def connect_data(self):
        self.conn = psycopg2.connect(
               database="current_weather",
               user='postgres',
               password='deva120808',
               host='localhost', port='5432')
        self.conn.autocommit = True
        self.cursor = self.conn.cursor()
        self.conn.autocommit = True
        self.cursor = self.conn.cursor()
 def temp_data(self):
        self.cursor.execute('drop table if exists temp_final')
        sql = '''CREATE TABLE temp_final(sys_id int ,
temp double precision ,temp_max double precision,temp_min double precision,feels_like double precision);'''
        self.cursor.execute(sql)
        for index,row in self.df1.iterrows():
            self.cursor.execute("insert into temp_final(sys_id,temp,temp_max,temp_min,feels_like) values (%s,%s,%s,%s,%s)",
                                (row.sys_id,row.temp,row.temp_max,row.temp_min,row.feels_like))
        self.conn.commit()
        sql1='''select * from temp_final;'''
        self.cursor.execute(sql1)
        self.conn.commit()
 def time_data(self):
        self.cursor = self.conn.cursor()
        self.cursor.execute('drop table if exists time_final')
        sql = '''CREATE TABLE date_final(sys_id int,
                                Date DATE NOT NULL,
                                Sunrise TIME,
                                Sunset time);'''
        self.cursor.execute(sql)
        for index,row in self.df1.iterrows():
            self.cursor.execute("insert into date_final (sys_id,Date,Sunrise,Sunset) values (%s,%s,%s,%s)",
                                (row.sys_id,row.Date,row.sunrise_time,row.sunset_time))
        self.conn.commit()
        sql1 = '''select * from date_final;'''
        self.cursor.execute(sql1)
        for i in self.cursor.fetchall():
               print(i)
        self.conn.commit()
 def weather_data(self):
        self.cursor = self.conn.cursor()
        self.cursor.execute('drop table if exists weather_final')
        sql = '''CREATE TABLE weather_final(sys_id int,
                                weather_id int,
                                main char(20),
                                icon varchar(20),
                                description varchar(50),
                                wind_speed double precision,
                                pressure int,
                                visibility int,
                                humidity int,
                                wind_deg int);'''
        self.cursor.execute(sql)
        for index,row in self.df1.iterrows():
            self.cursor.execute("insert into weather_final (sys_id,weather_id,main,icon,description,wind_speed,pressure,visibility,humidity,wind_deg) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                                (row.sys_id,row.weather_id,row.main,row.icon,row.description,row.wind_speed,row.pressure,row.visibility,row.humidity,row.wind_deg))
        self.conn.commit()
        sql1 = '''select * from weather_final;'''
        self.cursor.execute(sql1)
        for i in self.cursor.fetchall():
               print(i)
        self.conn.commit()
 def country_data(self):
        self.cursor = self.conn.cursor()
        self.cursor.execute('drop table if exists country_final')
        sql = '''CREATE TABLE country_final(city_id SERIAL PRIMARY KEY,
                                country varchar(20),
                                city_name char(50),
                                lon double precision,
                                lat double precision,
                                sys_id int);'''
        self.cursor.execute(sql)
        for index,row in self.df1.iterrows():
            self.cursor.execute("insert into country_final (city_id,country,city_name,lon,lat,sys_id) values (%s,%s,%s,%s,%s,%s)",
                                (row.city_id,row.country,row.city_name,row.lon,row.lat,row.sys_id))
        self.conn.commit()
        sql1 = '''select * from country_final;'''
        self.cursor.execute(sql1)
        self.conn.commit()
        self.conn.close()

o=etl()
o.get_data()
o.connect_data()
o.temp_data()
o.time_data()
o.weather_data()
o.country_data()





